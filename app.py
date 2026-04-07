from flask import Flask, render_template, request, redirect, session, send_file, jsonify
import sqlite3
from datetime import datetime, date
import os
import io
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import threading

# Certificate imports
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.platypus import Table, TableStyle

app = Flask(__name__)
app.secret_key = 'beyondclass'

DB = "beyond_class.db"

# ─────────────────────────────────────────────
#  EMAIL CONFIGURATION  ← UPDATE THESE VALUES
# ─────────────────────────────────────────────
EMAIL_SENDER   = "ashishkailskirde@gmail.com"        # Your Gmail address
EMAIL_PASSWORD = "oiejyabuzvazvith"      # Gmail App Password (16-char)
EMAIL_HOST     = "smtp.gmail.com"
EMAIL_PORT     = 587


def send_email_async(subject, html_body, recipient_emails):
    """Send emails in a background thread so the web request isn't blocked."""
    def _send():
        try:
            server = smtplib.SMTP(EMAIL_HOST, EMAIL_PORT)
            server.ehlo()
            server.starttls()
            server.login(EMAIL_SENDER, EMAIL_PASSWORD)

            for recipient in recipient_emails:
                msg = MIMEMultipart("alternative")
                msg["Subject"] = subject
                msg["From"]    = f"Beyond Class <{EMAIL_SENDER}>"
                msg["To"]      = recipient
                msg.attach(MIMEText(html_body, "html"))
                server.sendmail(EMAIL_SENDER, recipient, msg.as_string())

            server.quit()
            print(f"✅ Emails sent to {len(recipient_emails)} student(s).")
        except Exception as e:
            print(f"❌ Email error: {e}")

    thread = threading.Thread(target=_send, daemon=True)
    thread.start()


def build_activity_email(teacher_name, title, description, event_date, category, priority):
    """Build a styled HTML email body for a new activity notification."""
    priority_color = {
        "High":   "#e53935",
        "Medium": "#fb8c00",
        "Low":    "#43a047",
    }.get(priority, "#2a5298")

    return f"""
    <!DOCTYPE html>
    <html>
    <body style="margin:0;padding:0;background:#f0f4f8;font-family:'Segoe UI',sans-serif;">
      <table width="100%" cellpadding="0" cellspacing="0" bgcolor="#f0f4f8">
        <tr><td align="center" style="padding:40px 10px;">

          <table width="600" cellpadding="0" cellspacing="0"
                 style="background:#ffffff;border-radius:16px;
                        box-shadow:0 8px 30px rgba(0,0,0,0.12);overflow:hidden;">

            <!-- Header -->
            <tr>
              <td style="background:linear-gradient(135deg,#1e3c72,#2a5298);
                         padding:36px 40px;text-align:center;">
                <h1 style="color:#ffffff;margin:0;font-size:26px;letter-spacing:1px;">
                  🎓 Beyond Class
                </h1>
                <p style="color:rgba(255,255,255,0.8);margin:6px 0 0;font-size:14px;">
                  Activity Management System
                </p>
              </td>
            </tr>

            <!-- Badge -->
            <tr>
              <td style="background:#38ef7d;padding:10px 40px;text-align:center;">
                <span style="color:#1a1a1a;font-weight:700;font-size:13px;
                             letter-spacing:1.5px;text-transform:uppercase;">
                  🔔 New Activity Posted
                </span>
              </td>
            </tr>

            <!-- Body -->
            <tr>
              <td style="padding:36px 40px;">
                <p style="margin:0 0 20px;font-size:16px;color:#333;">
                  Hi there! A new activity has been posted by
                  <strong>{teacher_name}</strong>. Check it out below:
                </p>

                <!-- Activity Card -->
                <table width="100%" cellpadding="0" cellspacing="0"
                       style="background:#f8faff;border-radius:12px;
                              border:1px solid #dde6f5;margin-bottom:24px;">
                  <tr>
                    <td style="padding:28px 30px;">
                      <h2 style="margin:0 0 12px;color:#1e3c72;font-size:22px;">
                        {title}
                      </h2>
                      <p style="margin:0 0 20px;color:#555;font-size:15px;
                                line-height:1.6;">
                        {description}
                      </p>
                      <table cellpadding="0" cellspacing="0">
                        <tr>
                          <td style="padding:4px 14px;background:#e8f0fe;
                                     border-radius:20px;color:#2a5298;
                                     font-size:13px;font-weight:600;
                                     margin-right:8px;">
                            📅 {event_date}
                          </td>
                          <td width="10"></td>
                          <td style="padding:4px 14px;background:#e8f0fe;
                                     border-radius:20px;color:#2a5298;
                                     font-size:13px;font-weight:600;">
                            🏷️ {category}
                          </td>
                          <td width="10"></td>
                          <td style="padding:4px 14px;
                                     background:{priority_color};
                                     border-radius:20px;color:#fff;
                                     font-size:13px;font-weight:600;">
                            ⚡ {priority} Priority
                          </td>
                        </tr>
                      </table>
                    </td>
                  </tr>
                </table>

                <p style="margin:0 0 24px;font-size:15px;color:#444;">
                  Log in to your student dashboard to register for this activity
                  before it fills up!
                </p>

                <div style="text-align:center;">
                  <a href="#"
                     style="display:inline-block;padding:14px 38px;
                            background:linear-gradient(45deg,#11998e,#38ef7d);
                            color:#fff;font-weight:700;font-size:15px;
                            border-radius:30px;text-decoration:none;
                            box-shadow:0 6px 20px rgba(56,239,125,0.4);">
                    View &amp; Register
                  </a>
                </div>
              </td>
            </tr>

            <!-- Footer -->
            <tr>
              <td style="background:#f0f4f8;padding:20px 40px;
                         text-align:center;border-top:1px solid #dde6f5;">
                <p style="margin:0;color:#999;font-size:12px;">
                  This is an automated notification from
                  <strong>Beyond Class Activity Management System</strong>.<br>
                  Please do not reply to this email.
                </p>
              </td>
            </tr>

          </table>
        </td></tr>
      </table>
    </body>
    </html>
    """


# ---------------- DATABASE INIT ----------------

def init_db():
    conn = sqlite3.connect(DB)
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        email TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        role TEXT NOT NULL,
        first_login INTEGER DEFAULT 0
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS activities (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT NOT NULL,
        description TEXT,
        event_date TEXT,
        category TEXT,
        priority TEXT,
        teacher_id INTEGER,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (teacher_id) REFERENCES users(id)
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS registrations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        activity_id INTEGER,
        registered_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(user_id, activity_id),
        FOREIGN KEY (user_id) REFERENCES users(id),
        FOREIGN KEY (activity_id) REFERENCES activities(id)
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS notifications (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        activity_id INTEGER,
        message TEXT,
        is_read INTEGER DEFAULT 0,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (user_id) REFERENCES users(id),
        FOREIGN KEY (activity_id) REFERENCES activities(id)
    )
    """)

    cur.execute("SELECT id FROM users WHERE role='admin'")
    admin = cur.fetchone()

    if admin is None:
        cur.execute("""
        INSERT INTO users (name, email, password, role, first_login)
        VALUES (?, ?, ?, ?, ?)
        """, ("Admin", "admin@gmail.com", "admin123", "admin", 0))
        print("✅ Default Admin Created")
        print("Email: admin@gmail.com")
        print("Password: admin123")

    conn.commit()
    conn.close()


init_db()

# ---------------- CERTIFICATE GENERATION ----------------
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.lib import colors

def generate_certificate(student_name, activity_title, event_date):
    folder = "static/certificates"
    if not os.path.exists(folder):
        os.makedirs(folder)

    safe_name = student_name.replace(" ", "_")
    safe_title = activity_title.replace(" ", "_")
    filename = f"{safe_name}_{safe_title}.pdf"
    filepath = os.path.join(folder, filename)

    c = canvas.Canvas(filepath, pagesize=A4)
    width, height = A4

    c.setStrokeColorRGB(0.83, 0.68, 0.21)
    c.setLineWidth(10)
    c.rect(30, 30, width-60, height-60)
    c.setLineWidth(3)
    c.rect(50, 50, width-100, height-100)

    logo_path = os.path.join("static", "logo.png")
    if os.path.exists(logo_path):
        c.drawImage(logo_path, width/2 - 60, height - 160, width=120, height=120)

    c.setFont("Helvetica-Bold", 32)
    c.setFillColorRGB(0.6, 0.47, 0.16)
    c.drawCentredString(width/2, height - 220, "CERTIFICATE")
    c.setFont("Helvetica", 20)
    c.drawCentredString(width/2, height - 250, "OF PARTICIPATION")

    c.setFont("Helvetica", 16)
    c.setFillColor(colors.black)
    c.drawCentredString(width/2, height - 300, "This certificate is proudly presented to")

    c.setFont("Helvetica-Bold", 26)
    c.setFillColor(colors.darkred)
    c.drawCentredString(width/2, height - 340, student_name)

    c.setFont("Helvetica", 18)
    c.setFillColor(colors.black)
    c.drawCentredString(width/2, height - 380, f"For successfully participating in")

    c.setFont("Helvetica-Bold", 20)
    c.drawCentredString(width/2, height - 410, activity_title)

    c.setFont("Helvetica", 16)
    c.drawCentredString(width/2, height - 440, f"Held on {event_date}")

    sign_path = os.path.join("static", "signature.png")
    if os.path.exists(sign_path):
        c.drawImage(sign_path, width - 250, 140, width=150, height=60)

    c.setFont("Helvetica", 12)
    c.drawString(width - 230, 120, "Authorized Signature")

    c.setFont("Helvetica-Oblique", 10)
    c.drawCentredString(width/2, 80, "Beyond Class Activity Management System")

    c.save()
    return filepath


# ---------------- DOWNLOAD CERTIFICATE ----------------
@app.route('/download_certificate/<int:activity_id>')
def download_certificate(activity_id):
    if session.get('role') != 'student':
        return redirect('/')

    conn = sqlite3.connect(DB)
    cur = conn.cursor()

    cur.execute("""
        SELECT a.title, a.event_date, u.name
        FROM registrations r
        JOIN activities a ON r.activity_id = a.id
        JOIN users u ON r.user_id = u.id
        WHERE r.user_id=? AND r.activity_id=?
    """, (session['user_id'], activity_id))

    data = cur.fetchone()
    conn.close()

    if not data:
        return "You are not registered for this activity!"

    activity_title, event_date, student_name = data
    file_path = generate_certificate(student_name, activity_title, event_date)
    return send_file(file_path, as_attachment=True)


# ---------------- DOWNLOAD ACTIVITY REPORT (ADMIN) ----------------

@app.route('/download_activity_report/<int:activity_id>')
def download_activity_report(activity_id):

    if session.get('role') not in ['admin', 'teacher']:
        return redirect('/')

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    conn = sqlite3.connect(DB)
    cur = conn.cursor()

    if session.get('role') == 'teacher':
        cur.execute("SELECT teacher_id FROM activities WHERE id=?", (activity_id,))
        owner = cur.fetchone()
        if not owner or owner[0] != session['user_id']:
            conn.close()
            return "You cannot download a report for this activity!"

    cur.execute("""
        SELECT a.title, a.description, a.event_date, a.category, a.priority, u.name
        FROM activities a
        LEFT JOIN users u ON a.teacher_id = u.id
        WHERE a.id = ?
    """, (activity_id,))
    activity = cur.fetchone()

    if not activity:
        conn.close()
        return "Activity not found!"

    activity_title, activity_desc, event_date, category, priority, teacher_name = activity

    cur.execute("""
        SELECT u.name, u.email, r.registered_at
        FROM registrations r
        JOIN users u ON r.user_id = u.id
        WHERE r.activity_id = ?
        ORDER BY r.registered_at ASC
    """, (activity_id,))
    students = cur.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Activity Report"

    gold_fill   = PatternFill("solid", start_color="F0C040", end_color="F0C040")
    dark_fill   = PatternFill("solid", start_color="1F2D3D", end_color="1F2D3D")
    header_fill = PatternFill("solid", start_color="2E86AB", end_color="2E86AB")
    alt_fill    = PatternFill("solid", start_color="EAF4FB", end_color="EAF4FB")
    white_fill  = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")

    thin   = Side(style='thin', color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, w in zip(['A','B','C','D','E'], [6, 30, 32, 22, 16]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "BEYOND CLASS – Activity Registration Report"
    c.font = Font(name="Arial", bold=True, size=16, color="FFFFFF")
    c.fill = dark_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:E2")
    ws["A2"].fill = gold_fill
    ws.row_dimensions[2].height = 6

    info_rows = [
        ("Activity Title", activity_title),
        ("Description",    activity_desc or "—"),
        ("Event Date",     event_date or "—"),
        ("Category",       category or "—"),
        ("Priority",       priority or "—"),
        ("Created By",     teacher_name or "Admin"),
    ]

    for i, (label, value) in enumerate(info_rows, start=3):
        ws.merge_cells(f"B{i}:C{i}")
        ws.merge_cells(f"D{i}:E{i}")
        lc = ws.cell(row=i, column=2, value=label)
        vc = ws.cell(row=i, column=4, value=value)
        lc.font = Font(name="Arial", bold=True, size=11, color="1F2D3D")
        vc.font = Font(name="Arial", size=11, color="333333")
        lc.alignment = Alignment(vertical="center")
        vc.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[i].height = 20

    ws.merge_cells("A9:E9")
    ws["A9"].fill = gold_fill
    ws.row_dimensions[9].height = 6

    ws.merge_cells("A10:E10")
    cc = ws["A10"]
    cc.value = f"Total Registered Students: {len(students)}"
    cc.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    cc.fill = header_fill
    cc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[10].height = 26

    for col_idx, h in enumerate(["#", "Student Name", "Email Address", "Registered On", "Status"], 1):
        cell = ws.cell(row=11, column=col_idx, value=h)
        cell.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        cell.fill      = dark_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
    ws.row_dimensions[11].height = 22

    if students:
        for idx, (name, email, registered_at) in enumerate(students, start=1):
            row  = 11 + idx
            fill = alt_fill if idx % 2 == 0 else white_fill
            try:
                reg_date = registered_at.split(" ")[0] if registered_at else "—"
            except Exception:
                reg_date = registered_at or "—"

            for col_idx, val in enumerate([idx, name, email, reg_date, "✔ Registered"], 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.fill  = fill
                cell.font  = Font(
                    name="Arial", size=11,
                    color="2E7D32" if col_idx == 5 else "333333",
                    bold=(col_idx == 5)
                )
                cell.alignment = Alignment(
                    horizontal="center" if col_idx in [1, 4, 5] else "left",
                    vertical="center"
                )
                cell.border = border
            ws.row_dimensions[row].height = 20
    else:
        ws.merge_cells("A12:E12")
        nc = ws.cell(row=12, column=1, value="No students have registered for this activity yet.")
        nc.font      = Font(name="Arial", italic=True, size=11, color="999999")
        nc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[12].height = 24

    footer_row = 12 + max(len(students), 1) + 1
    ws.merge_cells(f"A{footer_row}:E{footer_row}")
    fc = ws.cell(row=footer_row, column=1,
                 value="Generated by Beyond Class Activity Management System")
    fc.font      = Font(name="Arial", italic=True, size=9, color="999999")
    fc.alignment = Alignment(horizontal="center")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_title = activity_title.replace(" ", "_").replace("/", "-")
    filename   = f"Report_{safe_title}.xlsx"

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ---------------- LOGIN ----------------
@app.route('/', methods=['GET','POST'])
def login():
    error = None

    if request.method == 'POST':
        email    = request.form['email']
        password = request.form['password']

        conn = sqlite3.connect(DB)
        cur  = conn.cursor()
        cur.execute("""
            SELECT id, name, role, first_login
            FROM users
            WHERE email=? AND password=?
        """, (email, password))
        user = cur.fetchone()
        conn.close()

        if user:
            session['user_id']  = user[0]
            session['username'] = user[1]
            session['role']     = user[2]

            if user[2] == 'student' and user[3] == 1:
                return redirect('/change_password')

            if user[2] == 'admin':
                return redirect('/admin')
            elif user[2] == 'teacher':
                return redirect('/teacher')
            else:
                return redirect('/student')
        else:
            error = "Invalid Email or Password. Please try again."

    return render_template("login.html", error=error)


# ---------------- FORGOT PASSWORD ----------------
@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        name  = request.form['name'].strip()
        email = request.form['email'].strip()

        conn = sqlite3.connect(DB)
        cur  = conn.cursor()
        cur.execute(
            "SELECT id FROM users WHERE LOWER(name)=LOWER(?) AND LOWER(email)=LOWER(?)",
            (name, email)
        )
        user = cur.fetchone()
        conn.close()

        if user:
            return redirect(f"/reset_password/{email}")
        else:
            return render_template("forgot_password.html",
                                   msg="No account found with that name and email combination.")

    return render_template("forgot_password.html")

# ---------------- RESET PASSWORD ----------------
@app.route('/reset_password/<email>', methods=['GET', 'POST'])
def reset_password(email):
    if request.method == 'POST':
        password = request.form.get('password', '').strip()
        confirm  = request.form.get('confirm', '').strip()

        if not password or not confirm:
            return render_template("reset_password.html", email=email, error="Please fill in all fields.")

        if password != confirm:
            return render_template("reset_password.html", email=email, error="Passwords do not match.")

        conn = sqlite3.connect(DB)
        cur  = conn.cursor()
        cur.execute("UPDATE users SET password=? WHERE email=?", (password, email))
        conn.commit()
        conn.close()

        return render_template("reset_password.html", email=email, success=True)

    return render_template("reset_password.html", email=email)


# ---------------- CHANGE PASSWORD (FIRST LOGIN) ----------------
@app.route('/change_password', methods=['GET','POST'])
def change_password():
    if session.get('role') != 'student':
        return redirect('/')

    if request.method == 'POST':
        new_password     = request.form['new_password']
        confirm_password = request.form['confirm_password']

        if new_password != confirm_password:
            return render_template("change_password.html", error="Passwords do not match!")

        conn = sqlite3.connect(DB)
        cur  = conn.cursor()
        cur.execute("UPDATE users SET password=?, first_login=0 WHERE id=?",
                    (new_password, session['user_id']))
        conn.commit()
        conn.close()
        return redirect('/student')

    return render_template("change_password.html")


# ---------------- LOGOUT ----------------
@app.route('/logout')
def logout():
    session.clear()
    return redirect('/')


# ---------------- ADMIN DASHBOARD ----------------
@app.route('/admin')
def admin():
    if session.get('role') != 'admin':
        return redirect('/')

    from datetime import date
    today = date.today()
    month_start = today.replace(day=1).isoformat()

    conn = sqlite3.connect(DB)
    cur  = conn.cursor()

    cur.execute("SELECT COUNT(*) FROM users WHERE role='student'")
    total_students = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM users WHERE role='teacher'")
    total_teachers = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM activities")
    total_activities = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM registrations")
    total_registrations = cur.fetchone()[0]

    cur.execute("SELECT id,name,email FROM users WHERE role='student'")
    students = cur.fetchall()
    cur.execute("SELECT id,name,email FROM users WHERE role='teacher'")
    teachers = cur.fetchall()
    cur.execute("""
        SELECT a.id, a.title, a.event_date, a.category, a.priority,
               u.name, COUNT(r.id) as reg_count
        FROM activities a
        LEFT JOIN users u ON a.teacher_id = u.id
        LEFT JOIN registrations r ON a.id = r.activity_id
        GROUP BY a.id ORDER BY a.event_date ASC
    """)
    activities = cur.fetchall()
    cur.execute("""
        SELECT u.name, u.email, a.title FROM registrations r
        JOIN users u ON r.user_id = u.id
        JOIN activities a ON r.activity_id = a.id
    """)
    registrations = cur.fetchall()

    cur.execute("""
        SELECT a.title, COUNT(r.id)
        FROM activities a LEFT JOIN registrations r ON a.id = r.activity_id
        GROUP BY a.id ORDER BY COUNT(r.id) DESC
    """)
    all_reg_per_activity = cur.fetchall()

    cur.execute("""
        SELECT strftime('%Y-%m', created_at) as mo, COUNT(*)
        FROM activities GROUP BY mo ORDER BY mo ASC LIMIT 12
    """)
    all_activities_by_month = cur.fetchall()

    cur.execute("""
        SELECT a.category, COUNT(r.id) as reg_count
        FROM activities a JOIN registrations r ON a.id = r.activity_id
        WHERE a.category IS NOT NULL AND a.category != ''
        GROUP BY a.category ORDER BY reg_count DESC
    """)
    all_category_regs = cur.fetchall()

    cur.execute("""
        SELECT a.title, COUNT(r.id)
        FROM activities a
        LEFT JOIN registrations r ON a.id = r.activity_id AND date(r.registered_at) >= ?
        GROUP BY a.id HAVING COUNT(r.id) > 0
        ORDER BY COUNT(r.id) DESC
    """, (month_start,))
    month_reg_per_activity = cur.fetchall()

    cur.execute("""
        SELECT strftime('%d', created_at) as day, COUNT(*)
        FROM activities WHERE date(created_at) >= ?
        GROUP BY day ORDER BY day ASC
    """, (month_start,))
    month_activities_by_day = cur.fetchall()

    cur.execute("SELECT COUNT(*) FROM registrations WHERE date(registered_at) >= ?", (month_start,))
    month_total_registrations = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM activities WHERE date(created_at) >= ?", (month_start,))
    month_total_activities = cur.fetchone()[0]

    cur.execute("""
        SELECT a.category, COUNT(r.id) as reg_count
        FROM activities a JOIN registrations r ON a.id = r.activity_id
        WHERE date(r.registered_at) >= ?
          AND a.category IS NOT NULL AND a.category != ''
        GROUP BY a.category ORDER BY reg_count DESC
    """, (month_start,))
    month_category_regs = cur.fetchall()

    conn.close()

    return render_template("admin_dashboard.html",
        total_students=total_students, total_teachers=total_teachers,
        total_activities=total_activities, total_registrations=total_registrations,
        students=students, teachers=teachers,
        activities=activities, registrations=registrations,
        all_reg_per_activity=all_reg_per_activity,
        all_activities_by_month=all_activities_by_month,
        all_category_regs=all_category_regs,
        month_reg_per_activity=month_reg_per_activity,
        month_activities_by_day=month_activities_by_day,
        month_category_regs=month_category_regs,
        month_total_registrations=month_total_registrations,
        month_total_activities=month_total_activities,
        current_month=today.strftime("%B %Y"),
    )

# ---------------- TEACHER DASHBOARD ----------------

@app.route('/teacher')
def teacher():
    if session.get('role') != 'teacher':
        return redirect('/')

    teacher_id = session['user_id']
    conn = sqlite3.connect(DB)
    cur  = conn.cursor()
    cur.execute("""
        SELECT a.title, a.event_date, a.category, a.priority,
               COUNT(r.id) as reg_count,
               a.id
        FROM activities a
        LEFT JOIN registrations r ON a.id = r.activity_id
        WHERE a.teacher_id = ?
        GROUP BY a.id
        ORDER BY a.event_date ASC
    """, (teacher_id,))
    activities = cur.fetchall()

    cur.execute("""
        SELECT a.title, u.name, u.email
        FROM registrations r
        JOIN activities a ON r.activity_id = a.id
        JOIN users u ON r.user_id = u.id
        WHERE a.teacher_id = ?
        ORDER BY a.title
    """, (teacher_id,))
    registrations = cur.fetchall()

    cur.execute("SELECT COUNT(*) FROM activities WHERE teacher_id=?", (teacher_id,))
    total_activities = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM users WHERE role='student'")
    total_students = cur.fetchone()[0]

    cur.execute("""
        SELECT COUNT(*) FROM registrations r
        JOIN activities a ON r.activity_id = a.id
        WHERE a.teacher_id = ?
    """, (teacher_id,))
    total_registrations = cur.fetchone()[0]

    conn.close()

    return render_template("teacher_dashboard.html",
        activities=activities,
        registrations=registrations,
        username=session.get('username'),
        total_activities=total_activities,
        total_students=total_students,
        total_registrations=total_registrations
    )

# ---------------- STUDENT DASHBOARD ----------------
@app.route('/student')
def student():
    if session.get('role') != 'student':
        return redirect('/')

    student_id = session['user_id']
    today = date.today()

    conn = sqlite3.connect(DB)
    cur  = conn.cursor()

    cur.execute("""
        SELECT id, message, is_read, created_at
        FROM notifications
        WHERE user_id=?
        ORDER BY created_at DESC
        LIMIT 5
    """, (student_id,))
    notifications = cur.fetchall()

    cur.execute("SELECT * FROM activities ORDER BY event_date ASC")
    activities_raw = cur.fetchall()

    activities = []
    for act in activities_raw:
        try:
            event_date = datetime.strptime(act[3], "%Y-%m-%d").date()
        except Exception:
            event_date = today
        status = "Upcoming" if event_date >= today else "Over"
        activities.append(act + (status,))

    cur.execute("""
        SELECT a.*
        FROM activities a
        JOIN registrations r ON a.id = r.activity_id
        WHERE r.user_id=?
    """, (student_id,))
    registered_activities = cur.fetchall()
    registered_ids = [r[0] for r in registered_activities]

    conn.close()

    return render_template("student_dashboard.html",
        activities=activities,
        registered_activities=registered_activities,
        registered_ids=registered_ids,
        today=today,
        notifications=notifications
    )


# ---------------- GET UNREAD NOTIFICATIONS ----------------
@app.route("/get_notifications")
def get_notifications():
    if 'user_id' not in session:
        return jsonify([])

    conn = sqlite3.connect(DB)
    cur  = conn.cursor()
    cur.execute("""
        SELECT id, message, created_at
        FROM notifications
        WHERE user_id=? AND is_read=0
        ORDER BY created_at DESC
    """, (session['user_id'],))
    rows = cur.fetchall()
    conn.close()

    return jsonify([{"id": r[0], "message": r[1], "time": r[2]} for r in rows])


# ---------------- MARK AS READ ----------------
@app.route("/mark_all_notifications_read")
def mark_all_notifications_read():
    if 'user_id' not in session:
        return "Not logged in"

    conn = sqlite3.connect(DB)
    cur  = conn.cursor()
    cur.execute("UPDATE notifications SET is_read=1 WHERE user_id=?", (session['user_id'],))
    conn.commit()
    conn.close()
    return "OK"


# ---------------- NOTIFICATION HISTORY ----------------
@app.route("/notifications")
def notification_history():
    if 'user_id' not in session:
        return redirect('/')

    conn = sqlite3.connect(DB)
    cur  = conn.cursor()
    cur.execute("""
        SELECT message, created_at FROM notifications
        WHERE user_id=? ORDER BY created_at DESC
    """, (session['user_id'],))
    notifications = cur.fetchall()
    conn.close()

    return render_template("notification_history.html", notifications=notifications)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  CREATE ACTIVITY  —  with EMAIL NOTIFICATION
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
@app.route('/create_activity', methods=['GET','POST'])
def create_activity():
    if session.get('role') not in ['admin','teacher']:
        return redirect('/')

    if request.method == 'POST':
        title       = request.form['title']
        description = request.form['description']
        event_date  = request.form['event_date']
        category    = request.form['category']
        priority    = request.form['priority']

        selected_date = datetime.strptime(event_date, "%Y-%m-%d").date()
        if selected_date < date.today():
            return "Cannot create activity for past date."

        teacher_name = session.get('username', 'Teacher')

        conn = sqlite3.connect(DB)
        cur  = conn.cursor()

        cur.execute("""
        INSERT INTO activities (title, description, event_date, category, priority, teacher_id)
        VALUES (?, ?, ?, ?, ?, ?)
        """, (title, description, event_date, category, priority, session['user_id']))

        # Fetch all student IDs and emails in one query
        cur.execute("SELECT id, email FROM users WHERE role='student'")
        students = cur.fetchall()

        student_emails = []
        for student_id, student_email in students:
            # In-app notification
            cur.execute(
                "INSERT INTO notifications (user_id, message) VALUES (?, ?)",
                (student_id, f"New Activity Posted: {title} on {event_date}")
            )
            student_emails.append(student_email)

        conn.commit()
        conn.close()

        # ── Send email notifications in background ──
        if student_emails:
            subject    = f"📢 New Activity: {title} | Beyond Class"
            html_body  = build_activity_email(
                teacher_name, title, description, event_date, category, priority
            )
            send_email_async(subject, html_body, student_emails)

        return render_template("create_activity.html", today=date.today(), success=True)

    return render_template("create_activity.html", today=date.today())


# ---------------- EDIT ACTIVITY ----------------
@app.route('/edit_activity/<int:id>', methods=['GET','POST'])
def edit_activity(id):
    if session.get('role') not in ['admin','teacher']:
        return redirect('/')

    conn = sqlite3.connect(DB)
    cur  = conn.cursor()

    if session['role'] == 'teacher':
        cur.execute("SELECT teacher_id FROM activities WHERE id=?", (id,))
        owner = cur.fetchone()
        if not owner or owner[0] != session['user_id']:
            conn.close()
            return "You cannot edit this activity!"

    if request.method == 'POST':
        cur.execute("""
        UPDATE activities
        SET title=?, description=?, event_date=?, category=?, priority=?
        WHERE id=?
        """, (
            request.form['title'],
            request.form['description'],
            request.form['event_date'],
            request.form['category'],
            request.form['priority'],
            id
        ))
        conn.commit()

        cur.execute("SELECT * FROM activities WHERE id=?", (id,))
        activity = cur.fetchone()
        conn.close()

        return render_template("edit_activity.html", activity=activity, success=True)

    cur.execute("SELECT * FROM activities WHERE id=?", (id,))
    activity = cur.fetchone()
    conn.close()
    return render_template("edit_activity.html", activity=activity)


# ---------------- REGISTER ----------------
@app.route('/register/<int:activity_id>')
def register(activity_id):
    if session.get('role') != 'student':
        return redirect('/')

    conn = sqlite3.connect(DB)
    cur  = conn.cursor()
    try:
        cur.execute("INSERT INTO registrations (user_id, activity_id) VALUES (?,?)",
                    (session['user_id'], activity_id))
        conn.commit()
    except Exception:
        pass
    conn.close()
    return redirect('/student')


# ---------------- DELETE ACTIVITY ----------------
@app.route('/delete_activity/<int:id>')
def delete_activity(id):
    if session.get('role') not in ['admin','teacher']:
        return redirect('/')

    conn = sqlite3.connect(DB)
    cur  = conn.cursor()
    cur.execute("DELETE FROM notifications WHERE activity_id=?", (id,))
    cur.execute("DELETE FROM registrations WHERE activity_id=?", (id,))
    cur.execute("DELETE FROM activities WHERE id=?", (id,))
    conn.commit()
    conn.close()

    return redirect('/teacher' if session.get('role') == 'teacher' else '/admin')


# ---------------- ADD STUDENT (TEACHER ONLY) ----------------
@app.route('/add_student', methods=['GET','POST'])
def add_student():
    if session.get('role') != 'teacher':
        return redirect('/')

    if request.method == 'POST':
        name     = request.form['name'].strip()
        email    = request.form['email'].strip()
        password = request.form['password'].strip()

        conn = sqlite3.connect(DB)
        cur  = conn.cursor()
        try:
            cur.execute("""
                INSERT INTO users (name, email, password, role, first_login)
                VALUES (?, ?, ?, 'student', 1)
            """, (name, email, password))
            conn.commit()
        except sqlite3.IntegrityError:
            conn.close()
            return render_template("add_student.html", error="Student with this email already exists!")
        conn.close()

        return render_template("add_student.html", success=True)

    return render_template("add_student.html")


# ---------------- ADD TEACHER (ADMIN ONLY) ----------------
@app.route('/add_teacher', methods=['GET','POST'])
def add_teacher():
    if session.get('role') != 'admin':
        return redirect('/')

    if request.method == 'POST':
        name     = request.form['name']
        email    = request.form['email']
        password = request.form['password']

        conn = sqlite3.connect(DB)
        cur  = conn.cursor()
        try:
            cur.execute("""
                INSERT INTO users (name, email, password, role)
                VALUES (?,?,?,'teacher')
            """, (name, email, password))
            conn.commit()
        except sqlite3.IntegrityError:
            conn.close()
            return render_template("add_teacher.html", error="Teacher with this email already exists!")
        conn.close()

        return render_template("add_teacher.html", success=True)

    return render_template("add_teacher.html")


if __name__ == '__main__':
    app.run(debug=True)